import calendar
from datetime import datetime
import openpyxl
import requests
import shutil
import sys

class Event:
               
    def __init__(self, rmtid, start_date, start_time, leq, sel, max_level, max_level_date_time, duration, operation_type, flight_number, aircraft_type, tail_number, beacon, other_port): 
        self.rmtid = rmtid
        self.start_date = start_date
        self.start_time = start_time
        self.leq = leq
        self.sel = sel
        self.max_level = max_level
        self.max_level_date_time = max_level_date_time
        self.duration = duration
        self.operation_type = operation_type
        self.flight_number = str(flight_number or "")
        self.aircraft_type = str(aircraft_type or "")
        self.tail_number = str(tail_number or "")
        self.beacon = beacon
        self.other_port = str(other_port or "")
            
    def __str__(self):
        return "{:>2}{:>10}{:>8}{:5.1f}{:5.1f}{:5.1f}{:>19}{:3d}{:>1}{:>7}{:>7}{:>8}{:>6}{:>4}".format(self.rmtid, self.start_date, self.start_time, self.leq, self.sel, self.max_level, self.max_level_date_time, self.duration, self.operation_type, self.flight_number, self.aircraft_type, self.tail_number, self.beacon, self.other_port)
                
class Summary:  
             
    def __init__(self, rmtid, month, uptime, totalleq, dnl, leq, events): 
        self.rmtid = rmtid
        self.month = month
        self.uptime = uptime
        self.totalleq = totalleq
        self.dnl = dnl
        self.leq = leq
        self.events = events        
        
    def __str__(self):
        return "{:>2}{:>3}{:5.1f}{:5.1f}{:5.1f}{:5.1f}{:6d}".format(self.rmtid, self.month, self.uptime, self.totalleq, self.dnl, self.leq, self.events)

def get_summary_data(year,month_num):
    url = 'http://public-reports-us-standard.s3-website-us-east-1.amazonaws.com/MWAA/{year}/{month_num:02} - {month_name}/Yearly Noise Summary (DCA) - {year} (Year To Date).xlsx'.format(year=year,month_num=month_num,month_name=calendar.month_name[month_num])
    
    local_filename = url.split('/')[-1]
    with requests.get(url, stream=True) as r:
        with open(local_filename, 'wb') as f:
            shutil.copyfileobj(r.raw, f)
    f.close 
    
    workbook=openpyxl.load_workbook(local_filename, read_only=True, data_only=True)
    workbook.iso_dates = True
    print(workbook.sheetnames)
    sheet=workbook.active

    RMTID=0
    MONTH=1
    UPTIME=2
    TOTALLEQ=3
    DNL=4
    LEQ=5
    EVENTS=6

    out = open('summary{year}.fw'.format(year=year), 'w') 
    
    for row in sheet.iter_rows(min_row=6,values_only=True):   
        if row[UPTIME] is not None:        
            summary =     Summary(row[RMTID],row[MONTH],row[UPTIME],row[TOTALLEQ] ,row[DNL], row[LEQ], row[EVENTS])
            out.write(summary.__str__() + '\n')                 
    out.close

def get_data(year,month_num,nmt):
    url = 'http://public-reports-us-standard.s3-website-us-east-1.amazonaws.com/MWAA/{year}/{month_num:02} - {month_name}/NMT {nmt:02} Noise Events - {year}-{month_name}.xlsx'.format(year=year,month_num=month_num,month_name=calendar.month_name[month_num],nmt=nmt)
          
    local_filename = url.split('/')[-1]
    with requests.get(url, stream=True) as r:
        with open(local_filename, 'wb') as f:
            shutil.copyfileobj(r.raw, f)
    f.close 
    
    RMTID=0
    START_DATE=1
    START_TIME=2
    LEQ=3
    SEL=4
    MAX_LEVEL=5
    MAX_LEVEL_DATE_TIME=6
    DURATION=7
    CLASSIFICATION=8
    OPERATION_TYPE=9
    FLIGHT_NUMBER=12
    AIRCRAFT_TYPE=13
    TAIL_NUMBER=14
    BEACON=15 
    OTHER_PORT=17
    
    try:
        workbook=openpyxl.load_workbook(local_filename,read_only=True, data_only=True)
        workbook.iso_dates = True
        sheet=workbook.active
        out = open('nmt{nmt:02}-{year}{month_num:02}.fw'.format( year=year, month_num=month_num, nmt=nmt),'w') 
        for row in sheet.iter_rows(min_row=4,values_only=True):
   
            if row[CLASSIFICATION] == 'Aircraft':
                start_date = row[START_DATE].strftime("%m/%d/%Y")
                start_time = row[START_TIME].strftime("%H:%M:%S")
                max_level_date_time =     row[MAX_LEVEL_DATE_TIME].strftime("%m/%d/%Y %H:%M:%S")
                event = Event(row[RMTID], start_date, start_time, row[LEQ], row[SEL], row[MAX_LEVEL],max_level_date_time,row[DURATION], row[OPERATION_TYPE],row[FLIGHT_NUMBER],row[AIRCRAFT_TYPE],row[TAIL_NUMBER],row[BEACON],row[OTHER_PORT])
       
                out.write(event.__str__() + '\n')                 
    
        out.close
    except Exception:
        print("ERROR: {0} is not a valid workbook".format(local_filename))

yyyymm=sys.argv[1]
print("\sys.argv[1]:", sys.argv[1])
year=int(sys.argv[1][0:4])
month_num = int(sys.argv[1][4:6])
print("year={0}, month_num={1}".format(year,month_num))

for nmt in (2,4,5,6,7,8,10,11):
    get_data(year,month_num,nmt)

get_summary_data(year,month_num)    
